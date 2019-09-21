import xlrd
import pdb
import os
import operator
from pprint import pprint as print
import datetime


#Development
from logbook import Logger, StreamHandler
import sys
StreamHandler(sys.stdout).push_application()
log = Logger('Logbook')
#log.warn('This is too cool for stdlib')

# current directory
dir_path = os.path.dirname(os.path.realpath(__file__))
filename = dir_path + '/schedule.xlsx'

import openpyxl
from openpyxl import load_workbook

import xlsxwriter
workbook = xlsxwriter.Workbook(filename)

import xlrd
book = xlrd.open_workbook(filename)
first_sheet = book.sheet_by_index(0)


# cell (row, column)

def find_bilingual_column():
    bilingual_column = None
    try:
        for i in range(0,25):
            if (first_sheet.cell(0,i).value == 'Bilingual'):
                bilingual_column = i
                #print(first_sheet.cell(0,i).value)
                #print(bilingual_column)
    except:
        pass
    return bilingual_column

bilingual_column = find_bilingual_column()

def read_first_sheet(first_sheet):
    schedule = list()
    count = 0
    for i in range(1,75):
        try:
            day = first_sheet.cell(i,1).value
            if len(day)> 1:
                count += 1
                value = {
                    'id': count,
                    'day': day,
                    'day_time': first_sheet.cell(i,2).value,
                    'hour': int(first_sheet.cell(i,3).value),
                    'school_1': first_sheet.cell(i,4).value,
                    'school_2': first_sheet.cell(i,5).value,
                    'school_3': first_sheet.cell(i,6).value,
                    'school_4': first_sheet.cell(i,7).value,
                    'school_5': first_sheet.cell(i,8).value,
                    'school_6': first_sheet.cell(i,9).value,
                    'school_7': first_sheet.cell(i,10).value,
                    'school_b': first_sheet.cell(i,bilingual_column).value,
                }
                schedule.append(value)
        except IndexError as exception:
            #print(exception.__class__.__name__ + " at line: " + str(i))
            pass
    return schedule

schedule = read_first_sheet(first_sheet)
schedule.sort(key=operator.itemgetter('id'))
# print(schedule)

###
# create excel file with xlsxwriter
###
workbook = xlsxwriter.Workbook('schedule_output.xlsx')



def find_school_color(school):
    if school == "Toronto":
        return "#4c9ad3"
    elif school == "Ottawa":
        return "#f58777"
    elif school == "Ryerson":
        return "#ffca09"
    elif school == "York":
        return "#df1f26"
    elif school == "Guelph":
        return "#f79027"
    elif school == "Brock":
        return "#7ecb4a"
    elif school == "Western":
        return "#765faa"
    elif school == "Algoma":
        return "#5c0e41"
    elif school == "Carleton":
        return "#1b8b48"
    elif school == "Lakehead":
        return "#7acfdb"
    elif school == "UOIT":
        return "#30d6b1"
    elif school == "OTECH":
        return "#cf5029"
    elif school == "Laurentian":
        return "#75b390"
    elif school == "G-Humber":
        return "#a65fa9"
    elif "umber" in school:
        return "#a65fa9"
    elif school == "McMaster":
        return "#7e2818"  
    elif school == "st-paul":
        return "#3e3e3d" 
    elif school == "Queens":
        return "#9f1f63" 
    elif school == "":
        return "#FFFFFF"    
    else:
        return '#a1a3a6'

def prepare_sheet(worksheet):
    for count_row in range(0,20):
        for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            sheet_format = workbook.add_format()
            sheet_format.set_bg_color('#FFFFFF')
            sheet_format.set_border(0)
            worksheet.write(column+ str(count_row) , "", sheet_format )


def find_school_format(school):
    school_format = workbook.add_format()
    #function call
    school_color = find_school_color(school)
    school_format.set_bg_color(school_color)

    school_format.set_font_size(20)
    school_format.set_align('center')
    if len(school) > 3:
        school_format.set_border(1)
    #school_format.set_center_across()
    school_format.set_font_color('#FFFFFF')
    return school_format

def change_hour_text_content(hour):
    if hour == 10:
        return '10:00 - 11:00am'
    if hour == 11:
        return '11:00 - 12:00pm'
    if hour == 12:
        return '12:00 - 1:00pm'
    if hour == 13:
        return '1:00 - 2:00pm'
    if hour == 14:
        return '2:00 - 3:00pm'
    if hour == 15:
        return '3:00 - 4:00pm'
    if hour == 16:
        return '4:00 - 5:00pm'
    if hour == 17:
        return '5:00 - 6:00pm'
    if hour == 18:
        return '6:00 - 7:00pm'
    if hour == 19:
        return '7:00 - 8:00pm'
    if hour == 20:
        return '8:00 - 9:00pm'
    if hour == 21:
        return '9:00 - 10:00pm'
    else:
        return 'unknown'

def change_hour_format(hour):
    hour = change_hour_text_content(hour)
    hour_format = workbook.add_format()
    hour_format.set_bg_color('#39393e')
    hour_format.set_font_color('#FFFFFF')
    hour_format.set_font_size(20)
    hour_format.set_align('center')
    return hour_format


def filter_by_day(day, schedule):
    given_day_schedule = [shift for shift in schedule if shift.get('day') == day]
    given_day_schedule.sort(key=operator.itemgetter('id'))
    return given_day_schedule


def write_schedule(day, day_schedule):
    count_row = 3
    worksheet = workbook.add_worksheet(day)
    prepare_sheet(worksheet)
    format_day_name = workbook.add_format()
    format_day_name.set_font_size(36)
    format_day_name.set_bold()
    format_day_name.set_align('left')
    # Write some data headers.
    worksheet.write('B1', day, format_day_name)
    # format Bilingual header
    format_bilingual_header= workbook.add_format()
    format_bilingual_header.set_font_size(22)
    format_bilingual_header.set_bold()
    format_bilingual_header.set_align('center')
    worksheet.write('J2', 'Bilingual', format_bilingual_header)
    worksheet.set_column(1, 12, 25)
    for shift in day_schedule:
        #print(shift)
        #worksheet.write('A'+ str(count_row) , shift.get('day') )
        worksheet.write('B'+ str(count_row) , change_hour_text_content(shift.get('hour')), change_hour_format(shift.get('hour')) )
        worksheet.write('C'+ str(count_row) , shift.get('school_1'), find_school_format(shift.get('school_1')) )
        worksheet.write('D'+ str(count_row) , shift.get('school_2'),  find_school_format(shift.get('school_2')) )
        worksheet.write('E'+ str(count_row) , shift.get('school_3'),  find_school_format(shift.get('school_3')) )
        worksheet.write('F'+ str(count_row) , shift.get('school_4'),  find_school_format(shift.get('school_4')) )
        worksheet.write('G'+ str(count_row) , shift.get('school_5'),  find_school_format(shift.get('school_5')) )
        worksheet.write('H'+ str(count_row) , shift.get('school_6'),  find_school_format(shift.get('school_6')) )
        worksheet.write('I'+ str(count_row) , shift.get('school_7'),  find_school_format(shift.get('school_7')) )
        worksheet.write('J'+ str(count_row) , shift.get('school_b'),  find_school_format(shift.get('school_b'))  )
        count_row +=1

monday_schedule = filter_by_day("Monday", schedule)
tuesday_schedule =  filter_by_day("Tuesday", schedule)
wednesday_schedule =  filter_by_day("Wednesday", schedule)
thursday_schedule =  filter_by_day("Thursday", schedule)
friday_schedule =  filter_by_day("Friday", schedule)
saturday_schedule = filter_by_day("Saturday", schedule)
sunday_schedule =  filter_by_day("Sunday", schedule)

write_schedule("Monday", monday_schedule)
write_schedule("Tuesday", tuesday_schedule)
write_schedule("Wednesday", wednesday_schedule)
write_schedule("Thursday", thursday_schedule)
write_schedule("Friday", friday_schedule)
write_schedule("Saturday", saturday_schedule)
write_schedule("Sunday", sunday_schedule)

#workbook.close()


column = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
sheet_name = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
#sheet_name = ['Monday', 'Tuesday',  'Wednesday']


def find_same_cell_value_as_previous():
    book = xlrd.open_workbook('schedule_output.xlsx')
    same_cell_value_as_previous_cell = list()
    for day in sheet_name:
            current_sheet = book.sheet_by_name(sheet_name=day)
            precedent = 'Uni'
            if day == 'Wednesday':
                import pdb
                #pdb.set_trace()
            for current_col in range(2, len(column)-1):
                for i in range(2,25):
                    try:
                        current = current_sheet.cell(i,current_col).value
                        if len(current) > 2:
                            if current == precedent:
                                ajustement = 0
                                cell_value = column[current_col] + str(i+1)+'--:' + current + "--:" + day
                                same_cell_value_as_previous_cell.append(cell_value)
                            precedent = current
                    except IndexError:
                        pass
                    except TypeError:
                        pass
    return same_cell_value_as_previous_cell


same_cell_value_as_previous_cell = find_same_cell_value_as_previous()

tuesday_only = [tuesday for tuesday in same_cell_value_as_previous_cell if 'uesday' in tuesday]
for cell in same_cell_value_as_previous_cell:
    location = cell.split('--:')[0]
    if 'C' in location:
        #print(cell)
        pass    

list_of_schools = [
"Toronto",
"Ottawa",
"Ryerson",
"York",
"Guelph",
"Brock",
"Western",
"Algoma",
"Carleton",
"Lakehead",
"UOIT",
"OTECH",
"Laurentian",
"G-Humber",
"McMaster",
"st-paul",
"Queens"
]


# removing Mentee name and other cells
duplicate = list()
for cell in same_cell_value_as_previous_cell:
    if 'ening' not in cell:
        if 'eekday' not in cell:
            if cell.split('--:')[1] in list_of_schools:
                #print(cell.split('--:'))
                duplicate.append(cell)


#print(duplicate)


monday_only = [monday for monday in duplicate if 'monday' in monday.lower()]
tuesday_only = [tuesday for tuesday in duplicate if 'tuesday' in tuesday.lower()]
wednesday_only = [wednesday for wednesday in duplicate if 'wednesday' in wednesday.lower()]
thursday_only = [thursday for thursday in duplicate if 'thursday' in thursday.lower()]
friday_only = [friday for friday in duplicate if 'friday' in friday.lower()]
saturday_only = [saturday for saturday in duplicate if 'saturday' in saturday.lower()]
sunday_only = [sunday for sunday in duplicate if 'sunday' in sunday.lower()]
print('\n\n\n')
print('-'*10)
print('-'*10)
#print(sunday_only)


def remove_following_cell(workbook, day, specific_day):
    worksheet =  workbook.get_worksheet_by_name(day)
    for cell in specific_day:
        location = cell.split('--:')[0]
        col = location[0]
        number = int(location[1::])
        school = cell.split('--:')[1]
        cell_format = workbook.add_format()
        if number < 15:
            try:
                worksheet.write(location, "", find_school_format(school))
            except:
                pass


day_only = [monday_only, tuesday_only, wednesday_only,
            thursday_only, friday_only, saturday_only, sunday_only]
day_name = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']


remove_following_cell(workbook, 'Monday', monday_only)
remove_following_cell(workbook, 'Tuesday', tuesday_only)
remove_following_cell(workbook, 'Wednesday', wednesday_only)
remove_following_cell(workbook, 'Thursday', thursday_only)
remove_following_cell(workbook, 'Friday', friday_only)
remove_following_cell(workbook, 'Saturday', saturday_only)
remove_following_cell(workbook, 'Sunday', sunday_only)

#230
workbook.close()


#from openpyxl
wb = load_workbook(filename = 'schedule_output.xlsx')



# or
for day in ['Friday', 'Saturday', 'Sunday']:
    sheet_name = wb[day]
    sheet_name.delete_cols(7,3)

wb.save(filename = dir_path + '/schedule_output.xlsx')



"""
Delete unused column --- do it manually instead

or
from openpyxl import load_workbook
wb = load_workbook(filename = 'schedule_output.xlsx')

#create manually a new sheet 
sheet_name = wb['Sheet1']
have_content = False
for col in ['F','G', 'H', 'I', 'J', 'K']:
    for row in range(1,5):
        cell = col+str(row)
        print(str(col)+str(row))
        content = sheet_name[cell].value
        print(content)
        if content != None:
            have_content = True
            print("Have content in " + col + " :: "+ content    )
            break
    if have_content == True:
        sheet_name.delete_cols(7,1)
    content = None
    have_content = False


###  TODO there is a bug in column

wb.save(filename = dir_path + '/schedule_output.xlsx')
"""